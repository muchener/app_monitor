# coding: utf-8
import requests
from bs4 import BeautifulSoup
import json
from shutil import copyfile
from openpyxl import load_workbook,Workbook
import time

class Doexcel:
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name        

    def get_data(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        test_data = []  
        for i in range(1,sheet.max_row):
            sub_data = {} 
            sub_data['no'] = sheet.cell(i+1,1).value
            sub_data['app_name'] = sheet.cell(i+1,2).value
            sub_data['channel'] = sheet.cell(i+1,3).value
            sub_data['package'] = sheet.cell(i+1,4).value
            sub_data['version'] = sheet.cell(i+1,5).value
            sub_data['update'] = sheet.cell(i+1,6).value
            test_data.append(sub_data)
        wb.close()
        return test_data  #返回获取到的数据

    # 版本、更新日期写入
    def write_data(self,data):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        for i in data:
            sheet.cell(row = i['no']+1,column=5,value=i['version'])
            sheet.cell(row = i['no']+1,column=6,value=i['update'])
        wb.save(self.file_name)
        

class Mon:

    def call_request(self, url, num=0):
        try:
            res = requests.get(url)
            return res
        except Exception as e:
            if num <= 3:
                num += 1
                self.call_request(url, num)
            else:
                print("Error ,", str(e))


    def send_alert(self,appname, channel, old_version, uptime,new_version,message="null"):
        # 这是告警方法
        print(" {} App名称 : {} 渠道 : {} {}-->{}".format(uptime, appname, channel,old_version, new_version))
        # 写到update文件里
        with open("update.txt", 'a+', encoding='utf-8') as f:
            data = "{} App名称 : {} 渠道 : {} {}-->{}\n".format(uptime,appname, channel,old_version, new_version)
            old = f.read()
            f.seek(0)
            f.write(data)
            f.write(old)

    # def call_tencent(self, package):
    #     url = 'https://sj.qq.com/myapp/detail.htm?apkName={}'.format(package)
    #     res = self.call_request(url)
    #     try:
    #         res = BeautifulSoup(res.text, 'lxml')
    #         res = res.find_all('div', class_='det-othinfo-data')
    #         version = res[0].text
    #         return version
    #     except Exception as e:
    #         print("未获取到版本信息")
    #         print(str(e))


    def call_xiaomi(self,package):
        url = 'https://app.mi.com/details?id={}'.format(package)
        res = self.call_request(url)
        try:
            res = BeautifulSoup(res.text, 'lxml')
            res = res.find_all('div', style='float:right;')
            version = res[2].text
            return version
        except Exception as e:
            print("未获取到版本信息")
            print(str(e))


    def call_huawei(self,package):
        url = 'https://web-drcn.hispace.dbankcloud.cn/uowap/index?method=internal.getTabDetail&serviceType=20&reqPageNum=1&maxResults=25&uri=app%7C{}&locale=zh'.format(package)
        res = self.call_request(url)
        try:
            res = json.loads(res.text)
            version = res["layoutData"][1]['dataList'][0]['versionName']
            return version
        except Exception as e:
            print("未获取到版本信息")
            print(str(e))


    def call_apple(self,package):
        url = "https://apps.apple.com/cn/app/{}".format(package)
        res = self.call_request(url)
        try:
            res = BeautifulSoup(res.text, 'lxml')
            version = res.find_all('p', class_='l-column small-6 medium-12 whats-new__latest__version')
            return version[0].text
        except Exception as e:
            print("未获取到版本信息")
            print(str(e))

    def job(self):
        copyfile("data.xlsx", "bak_data.xlsx")
        data = Doexcel("data.xlsx",'sheet1').get_data()
        new_data = []
        for i in data:
            no = i['no']
            channel_name = i['channel']
            package_name = i['package']
            version = i['version']
            app_name = i['app_name']
            uptime = i['update']
            func = getattr(self, "call_"+channel_name, None)
            new_version = func(package_name)
            if new_version != version:
                uptime = time.strftime("%Y-%m-%d", time.localtime())
                self.send_alert(app_name, channel_name, version, uptime,new_version)
                version = new_version
            new_data.append({"no":no,"channel": channel_name, "package": package_name, "version": new_version, "app_name": app_name,"update":uptime})
        Doexcel("data.xlsx",'sheet1').write_data(new_data)

if __name__ == '__main__':
    mon = Mon()
    mon.job()
